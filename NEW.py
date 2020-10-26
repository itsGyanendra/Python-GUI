import tkinter as tk
from functools import partial
import openpyxl


from PIL import ImageTk, Image

path = "/home/gyanendra/PycharmProjects/guiProject/yogaasan_data.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

max_row = ws.max_row

def call_result(label_result, n1, n2):
    num1 = (n1.get())
    num2 = (n2.get())
    value = ''
    for i in range(1, max_row + 1):
        cell_id = ws.cell(row=i, column=1)
        if (((ws["D" + str(i)].value) != None) and ((ws["B" + str(i)].value) != None)):
            if ((((ws["D" + str(i)].value).find(num1)) != -1) and (
            (((ws["B" + str(i)].value).find(num2)) != -1))):
                # print((ws["A"+str(i+1)].value)
                value = f'{value + str(cell_id.value)}\n'
        else:
            if(((ws["B" + str(i)].value) != None) and (((ws["B" + str(i)].value).find(num2)) != -1)):
                value = f'{value + str(cell_id.value)}\n'


    label_result.config(text=value)

    return


top= tk.Tk()
top.title('Yoga Chooser')
top.geometry("10000x4000")



img = ImageTk.PhotoImage(file="/home/gyanendra/PycharmProjects/guiProject/images.png")
scale_w = 1
scale_h = 1
#img = img.PhotoImage_photo.zoom(scale_w, scale_h)

background = tk.Label(top, image=img, bd=10)
background.pack(fill='both', expand=True)
background.image = img

# resize empty rows, columns to put other elements in center
#background.rowconfigure(0, weight=100)
#background.rowconfigure(10, weight=100)
#background.columnconfigure(0, weight=100)
#background.columnconfigure(10, weight=100)


number1 = tk.StringVar()
number2 = tk.StringVar()

label_name = tk.Label(background, bg="navajo white", text="Name").grid(row=1, column=0, padx=(25, 20), pady=(100, 5))
label_gender = tk.Label(background, bg="navajo white", text="Gender").grid(row=2, column=0, pady=(10, 0))
label_age = tk.Label(background, bg="navajo white", text="Age").grid(row=3, column=0, pady=(10, 0))
label_weight = tk.Label(background, bg="navajo white", text="Weight(in kg)").grid(row=4, column=0, pady=(10, 0))
label_height = tk.Label(background, bg="navajo white", text="Height( in m)").grid(row=5, column=0, pady=(10, 0))
label_bmi = tk.Label(background, bg="navajo white", text="BMI INDEX").grid(row=6, column=0, pady=(10, 0))
labelNum1 = tk.Label(background, bg="navajo white", text="Disease").grid(row=7, column=0, pady=(10, 0))
labelNum2 = tk.Label(background, bg="navajo white", text="Type of Yoga").grid(row=8, column=0, pady=(10, 0))

labelResult = tk.Label(background)

labelResult.grid(row=10, column=2, padx=(70, 20), pady=(10, 5))

name = tk.Entry(background, width=50).grid(row=1, column=2, padx=(70, 20), pady=(100, 5))
gender = tk.Entry(background, width=50).grid(row=2, column=2, padx=(70, 20), pady=(10, 0))
age = tk.Entry(background, width=50).grid(row=3, column=2, padx=(70, 20), pady=(10, 0))
weight = tk.Entry(background, width=50).grid(row=4, column=2, padx=(70, 20), pady=(10, 0))
height = tk.Entry(background, width=50).grid(row=5, column=2, padx=(70, 20), pady=(10, 0))
bmi = tk.Entry(background, width=50).grid(row=6, column=2, padx=(70, 20), pady=(10, 0))
entryNum1 = tk.Entry(background, textvariable=number1, width=50).grid(row=7, column=2, padx=(70, 20), pady=(10, 0))

entryNum2 = tk.Entry(background, textvariable=number2, width=50).grid(row=8, column=2, padx=(70, 20), pady=(10, 0))

call_result = partial(call_result, labelResult, number1, number2)

buttonCal = tk.Button(background, text="Show Results", command=call_result).grid(row=9, column=2, padx=(70, 20), pady=(10, 0))

top.mainloop()
